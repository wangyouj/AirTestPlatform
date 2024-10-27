'''
@author: No.47
@file: test.py
@time: 2024/8/30 16:13
@desc:
'''
from apscheduler.schedulers.background import BackgroundScheduler
# currentPage=3
# pageSize=20
# start = (currentPage - 1) * pageSize
# dataSql='select * from tt_atms_budget_import_test_data  order by id '
# dataSql = dataSql + ' limit {0},{1};'.format(start, pageSize)
# print(dataSql)

from openpyxl import load_workbook

from business.AirBudget import product_and_export_budget_data
from common.AirDecorators import AirIntervalScheduler
from common.Tools import bjtm

# def updateExcel():
#     '''
#
#     :return:
#     常见异常：row[5].value = row2[1].value IndexError: tuple index out of range
#     原因及处理方案：row[5]对应列列名为空导致，确认后补全列名即可
#     '''
#     filePath = 'D:/2024版本/预算管理/budget0815.xlsx'
#     # 加载Excel文件
#     workbook = load_workbook(filePath)
#
#     # 获取sheet1和sheet2
#     sheet1 = workbook['Sheet1']
#     sheet2 = workbook['Sheet2']
#
#     for row in sheet1.iter_rows():
#         value = row[0].value
#         print(value)
#         # 遍历sheet2的每一行
#         for row2 in sheet2.iter_rows():
#             # 如果sheet2当前行的A列值等于sheet1当前行的A列值
#             print(row2[0].value)
#             if value == row2[0].value:
#                 print('匹配OK!' + '归属网点=' + value + '----责任区=' + row2[1].value)
#                 # 将sheet2当前行的B列值填充到sheet1的K列
#                 row[5].value = row2[1].value
#
#     # 保存更新后的Excel文件
#     workbook.save(filePath)






































